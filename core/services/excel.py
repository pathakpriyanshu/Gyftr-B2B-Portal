"""Generates the .xlsx voucher delivery file — mirrors `src/lib/excel.ts`."""
import io
from datetime import datetime, date, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _indian_group(n) -> str:
    """Group an integer the Indian way: 12,34,567."""
    n = int(round(float(n)))
    sign = "-" if n < 0 else ""
    s = str(abs(n))
    if len(s) <= 3:
        return sign + s
    head, tail = s[:-3], s[-3:]
    parts = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    parts.insert(0, head)
    return sign + ",".join(parts) + "," + tail


def _fmt_date(value, with_time=False) -> str:
    if value is None:
        return "—"
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            try:
                value = datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                return value
    if isinstance(value, datetime):
        if value.tzinfo:
            value = value.astimezone(timezone.utc)
        out = value.strftime("%d %b %Y")
        if with_time:
            out += ", " + value.strftime("%I:%M %p").lower().lstrip("0")
        return out
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return str(value)


def generate_voucher_workbook(order, vouchers) -> bytes:
    wb = Workbook()
    wb.properties.creator = "Gyftr B2B Portal"
    wb.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)

    thin = Side(style="thin", color="FFE5E7EB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # --- Summary sheet ---
    summary = wb.active
    summary.title = "Summary"
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 40
    summary.append(["Field", "Value"])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1A2552")
    rows = [
        ("Order Number", order.order_number),
        ("Order Date", _fmt_date(order.created_at, True)),
        ("Total Vouchers", str(len(vouchers))),
        ("Total Face Value", f"INR {_indian_group(order.total_face_value)}"),
        ("Payment Status", order.payment_status),
        ("Generated On", _fmt_date(datetime.now(timezone.utc), True)),
    ]
    for field, value in rows:
        summary.append([field, value])

    # --- Vouchers sheet ---
    sheet = wb.create_sheet("Vouchers")
    headers = ["S.No", "Brand", "Denomination (INR)", "Voucher Code", "PIN", "Expiry Date"]
    widths = [8, 26, 20, 30, 14, 16]
    for i, w in enumerate(widths):
        sheet.column_dimensions[chr(ord("A") + i)].width = w
    sheet.append(headers)
    header_row = sheet[1]
    for cell in header_row:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFE6007E")
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22

    for i, v in enumerate(vouchers):
        sheet.append([
            i + 1,
            v.brand_name,
            v.denomination,
            v.code,
            v.pin or "—",
            _fmt_date(v.expiry_date) if v.expiry_date else "—",
        ])
        if i % 2 == 1:
            for cell in sheet[i + 2]:
                cell.fill = PatternFill("solid", fgColor="FFFDF2F8")

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
